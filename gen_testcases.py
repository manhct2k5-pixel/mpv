import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── colour palette ──────────────────────────────────────────────────────────
C_HEADER      = "1F4E79"   # deep blue  – header bg
C_HEADER_FONT = "FFFFFF"   # white      – header text
C_UD          = "D6E4F7"   # light blue – unauthenticated
C_KH          = "E2EFDA"   # light green – customer
C_NB          = "FFF2CC"   # yellow     – seller
C_NVK         = "FCE4D6"   # orange     – warehouse
C_QTV         = "E2CFEF"   # purple     – admin
C_ALT_UD      = "EBF3FB"
C_ALT_KH      = "F0F8ED"
C_ALT_NB      = "FFFAE6"
C_ALT_NVK     = "FDF0EA"
C_ALT_QTV     = "EEE5F6"
C_PRIORITY_H  = "FF0000"   # red   – High priority font
C_PRIORITY_M  = "FF8C00"   # orange
C_PRIORITY_L  = "2E75B6"   # blue

ACTOR_COLORS = {
    "Người dùng chưa đăng nhập": (C_UD,  C_ALT_UD),
    "Khách hàng":                 (C_KH,  C_ALT_KH),
    "Người bán":                  (C_NB,  C_ALT_NB),
    "Nhân viên kho":              (C_NVK, C_ALT_NVK),
    "Quản trị viên":              (C_QTV, C_ALT_QTV),
}

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def cell_style(ws, row, col, value, fill_hex, bold=False,
               wrap=True, align_h="left", align_v="top",
               font_color="000000", font_size=9):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=fill_hex)
    c.font = Font(bold=bold, color=font_color, size=font_size,
                  name="Calibri")
    c.alignment = Alignment(horizontal=align_h, vertical=align_v,
                             wrap_text=wrap)
    c.border = thin_border()
    return c

# ── test data ────────────────────────────────────────────────────────────────
HEADERS = [
    "STT", "Mã TC", "Tên Test Case", "Actor", "Chức năng",
    "Điều kiện tiên quyết", "Dữ liệu đầu vào",
    "Các bước thực hiện", "Kết quả mong đợi",
    "Kết quả thực tế", "Trạng thái", "Mức ưu tiên", "Loại test"
]

ROWS = [
    # ── Người dùng chưa đăng nhập ──────────────────────────────────────────
    (1,"TC-UD-001","Xem danh sách sản phẩm khi chưa đăng nhập","Người dùng chưa đăng nhập","Duyệt sản phẩm",
     "Chưa đăng nhập vào hệ thống","Không có",
     "1. Truy cập trang chủ\n2. Nhấn menu \"Sản phẩm\"",
     "Trang danh sách sản phẩm hiển thị đầy đủ, không yêu cầu đăng nhập","","","Cao","Functional"),
    (2,"TC-UD-002","Xem chi tiết sản phẩm khi chưa đăng nhập","Người dùng chưa đăng nhập","Xem sản phẩm",
     "Chưa đăng nhập","URL sản phẩm cụ thể",
     "1. Truy cập URL /san-pham/{slug} của bất kỳ sản phẩm nào",
     "Trang chi tiết hiển thị: tên, giá, hình ảnh, mô tả, đánh giá","","","Cao","Functional"),
    (3,"TC-UD-003","Thêm vào giỏ hàng khi chưa đăng nhập","Người dùng chưa đăng nhập","Giỏ hàng",
     "Chưa đăng nhập","Sản phẩm bất kỳ",
     "1. Truy cập trang sản phẩm\n2. Nhấn nút \"Thêm vào giỏ\"",
     "Hệ thống chuyển hướng đến trang đăng nhập hoặc hiển thị thông báo yêu cầu đăng nhập","","","Cao","Functional"),
    (4,"TC-UD-004","Truy cập trang thanh toán khi chưa đăng nhập","Người dùng chưa đăng nhập","Phân quyền",
     "Chưa đăng nhập","URL /thanh-toan",
     "1. Nhập trực tiếp URL /thanh-toan vào trình duyệt",
     "Hệ thống chuyển hướng về /login","","","Cao","Functional"),
    (5,"TC-UD-005","Truy cập trang Admin khi chưa đăng nhập","Người dùng chưa đăng nhập","Phân quyền",
     "Chưa đăng nhập","URL /admin",
     "1. Nhập trực tiếp URL /admin vào trình duyệt",
     "Hệ thống chuyển hướng về /login hoặc hiển thị 403 Forbidden","","","Cao","Security"),
    (6,"TC-UD-006","Tìm kiếm sản phẩm khi chưa đăng nhập","Người dùng chưa đăng nhập","Tìm kiếm",
     "Chưa đăng nhập","Từ khóa: \"áo\"",
     "1. Vào /san-pham\n2. Nhập từ khóa \"áo\" vào ô tìm kiếm\n3. Nhấn Enter",
     "Danh sách sản phẩm lọc theo từ khóa, hiển thị đúng kết quả","","","Trung bình","Functional"),
    (7,"TC-UD-007","Đăng ký tài khoản với thông tin hợp lệ","Người dùng chưa đăng nhập","Đăng ký",
     "Chưa có tài khoản, email chưa đăng ký",
     "Email: testuser01@gmail.com\nHọ tên: Nguyễn Văn A\nMật khẩu: Test@12345",
     "1. Truy cập /register\n2. Nhập đầy đủ thông tin hợp lệ\n3. Nhấn \"Đăng ký\"",
     "Tài khoản được tạo thành công, hệ thống chuyển về trang đăng nhập","","","Cao","Functional"),
    (8,"TC-UD-008","Đăng ký với email đã tồn tại","Người dùng chưa đăng nhập","Đăng ký",
     "Email đã được đăng ký trước đó",
     "Email: existing@gmail.com\nMật khẩu: Test@12345",
     "1. Truy cập /register\n2. Nhập email đã tồn tại\n3. Nhấn \"Đăng ký\"",
     "Hệ thống hiển thị lỗi: email đã được sử dụng","","","Cao","Functional"),
    (9,"TC-UD-009","Đăng ký với mật khẩu quá ngắn","Người dùng chưa đăng nhập","Đăng ký",
     "Chưa có tài khoản",
     "Email: newuser@gmail.com\nMật khẩu: 123",
     "1. Truy cập /register\n2. Nhập mật khẩu \"123\"\n3. Nhấn \"Đăng ký\"",
     "Hệ thống hiển thị lỗi validate mật khẩu không hợp lệ","","","Trung bình","Functional"),
    (10,"TC-UD-010","Xem lookbook khi chưa đăng nhập","Người dùng chưa đăng nhập","Xem nội dung",
     "Chưa đăng nhập","Không có",
     "1. Truy cập trang /lookbook",
     "Trang lookbook hiển thị đầy đủ không yêu cầu đăng nhập","","","Thấp","Functional"),

    # ── Khách hàng ──────────────────────────────────────────────────────────
    (11,"TC-KH-001","Đăng nhập với thông tin hợp lệ","Khách hàng","Đăng nhập",
     "Đã có tài khoản khách hàng hợp lệ",
     "Email: customer@gmail.com\nMật khẩu: đúng",
     "1. Truy cập /login\n2. Nhập email và mật khẩu đúng\n3. Nhấn \"Đăng nhập\"",
     "Đăng nhập thành công, chuyển hướng về trang chủ","","","Cao","Functional"),
    (12,"TC-KH-002","Đăng nhập sai mật khẩu","Khách hàng","Đăng nhập",
     "Đã có tài khoản",
     "Email: customer@gmail.com\nMật khẩu: saimatkhau123",
     "1. Truy cập /login\n2. Nhập email đúng, mật khẩu sai\n3. Nhấn \"Đăng nhập\"",
     "Hiển thị thông báo: sai email hoặc mật khẩu","","","Cao","Functional"),
    (13,"TC-KH-003","Đăng nhập với email không tồn tại","Khách hàng","Đăng nhập",
     "Không có tài khoản với email này",
     "Email: khongtontai@gmail.com\nMật khẩu: Test@123",
     "1. Truy cập /login\n2. Nhập email không tồn tại\n3. Nhấn \"Đăng nhập\"",
     "Hệ thống hiển thị thông báo lỗi phù hợp","","","Trung bình","Functional"),
    (14,"TC-KH-004","Thêm sản phẩm có hàng vào giỏ","Khách hàng","Giỏ hàng",
     "Đã đăng nhập; sản phẩm có tồn kho > 0",
     "Sản phẩm có size M còn hàng",
     "1. Đăng nhập\n2. Vào trang sản phẩm\n3. Chọn size M\n4. Nhấn \"Thêm vào giỏ\"",
     "Sản phẩm được thêm vào giỏ, số lượng giỏ tăng lên 1","","","Cao","Functional"),
    (15,"TC-KH-005","Thêm sản phẩm hết hàng vào giỏ","Khách hàng","Giỏ hàng",
     "Đã đăng nhập; sản phẩm tồn kho = 0",
     "Sản phẩm có size XL hết hàng",
     "1. Đăng nhập\n2. Vào trang sản phẩm\n3. Chọn size XL (hết hàng)\n4. Nhấn \"Thêm vào giỏ\"",
     "Nút thêm vào giỏ bị vô hiệu hóa hoặc hiển thị \"Hết hàng\"","","","Cao","Functional"),
    (16,"TC-KH-006","Cập nhật số lượng sản phẩm trong giỏ","Khách hàng","Giỏ hàng",
     "Đã đăng nhập; giỏ có ít nhất 1 sản phẩm",
     "Số lượng mới: 3",
     "1. Vào /gio-hang\n2. Thay đổi số lượng sản phẩm thành 3",
     "Số lượng được cập nhật, tổng tiền tự động tính lại","","","Trung bình","Functional"),
    (17,"TC-KH-007","Xóa sản phẩm khỏi giỏ hàng","Khách hàng","Giỏ hàng",
     "Đã đăng nhập; giỏ có sản phẩm","Không có",
     "1. Vào /gio-hang\n2. Nhấn nút xóa bên cạnh sản phẩm",
     "Sản phẩm bị xóa, tổng tiền cập nhật","","","Trung bình","Functional"),
    (18,"TC-KH-008","Áp mã giảm giá hợp lệ","Khách hàng","Mã giảm giá",
     "Đã đăng nhập; giỏ đủ điều kiện tối thiểu; mã còn hạn",
     "Mã: SALE10 (giảm 10%, đơn tối thiểu 200k)",
     "1. Vào /gio-hang\n2. Nhập mã SALE10\n3. Nhấn \"Áp dụng\"",
     "Mã áp dụng thành công, hiển thị số tiền giảm, tổng đơn cập nhật","","","Cao","Functional"),
    (19,"TC-KH-009","Áp mã giảm giá đã hết hạn","Khách hàng","Mã giảm giá",
     "Đã đăng nhập; mã hết hạn tồn tại trong hệ thống",
     "Mã: EXPIRED20",
     "1. Vào /gio-hang\n2. Nhập mã EXPIRED20\n3. Nhấn \"Áp dụng\"",
     "Hệ thống báo lỗi: \"Voucher đã hết hạn\"","","","Cao","Functional"),
    (20,"TC-KH-010","Áp mã giảm giá không tồn tại","Khách hàng","Mã giảm giá",
     "Đã đăng nhập","Mã: ABCXYZ999",
     "1. Vào /gio-hang\n2. Nhập mã ABCXYZ999\n3. Nhấn \"Áp dụng\"",
     "Hệ thống báo lỗi: \"Mã voucher không tồn tại\"","","","Cao","Functional"),
    (21,"TC-KH-011","Áp mã chưa đủ đơn tối thiểu","Khách hàng","Mã giảm giá",
     "Đã đăng nhập; giỏ < 200.000đ; mã yêu cầu tối thiểu 200k",
     "Mã: MIN200K | Tổng giỏ: 150.000đ",
     "1. Vào /gio-hang với tổng < 200k\n2. Nhập mã MIN200K\n3. Nhấn \"Áp dụng\"",
     "Hệ thống báo lỗi: \"Đơn hàng chưa đạt giá trị tối thiểu để áp voucher\"","","","Trung bình","Functional"),
    (22,"TC-KH-012","Áp mã tại trang thanh toán (Mua ngay)","Khách hàng","Mã giảm giá",
     "Đã đăng nhập; vào thanh toán qua nút Mua ngay; có mã hợp lệ",
     "Mã: SALE10",
     "1. Nhấn Mua ngay từ trang sản phẩm\n2. Tại /thanh-toan nhập mã SALE10\n3. Nhấn Áp dụng",
     "Mã được áp dụng ngay tại trang thanh toán, tổng tiền giảm","","","Cao","Functional"),
    (23,"TC-KH-013","Đặt hàng thành công bằng COD","Khách hàng","Đặt hàng",
     "Đã đăng nhập; giỏ có sản phẩm; địa chỉ hợp lệ",
     "Phương thức: COD",
     "1. Vào /gio-hang\n2. Nhấn \"Thanh toán\"\n3. Điền địa chỉ\n4. Chọn COD\n5. Nhấn \"Đặt hàng ngay\"",
     "Đơn hàng được tạo, chuyển đến trang xác nhận đơn","","","Cao","Functional"),
    (24,"TC-KH-014","Đặt hàng bằng chuyển khoản ngân hàng","Khách hàng","Đặt hàng",
     "Đã đăng nhập; giỏ có sản phẩm",
     "Phương thức: BANK_TRANSFER",
     "1. Vào /gio-hang\n2. Nhấn \"Thanh toán\"\n3. Điền địa chỉ\n4. Chọn Chuyển khoản\n5. Nhấn \"Đặt hàng ngay\"",
     "Đơn được tạo, hiển thị thông tin QR chuyển khoản","","","Cao","Functional"),
    (25,"TC-KH-015","Đặt hàng thiếu địa chỉ giao hàng","Khách hàng","Đặt hàng",
     "Đã đăng nhập; giỏ có sản phẩm",
     "Địa chỉ: để trống",
     "1. Vào /thanh-toan\n2. Để trống trường địa chỉ\n3. Nhấn \"Đặt hàng ngay\"",
     "Hệ thống hiển thị lỗi: yêu cầu điền địa chỉ","","","Cao","Functional"),
    (26,"TC-KH-016","Mua ngay từ trang sản phẩm","Khách hàng","Đặt hàng",
     "Đã đăng nhập; đã chọn size/màu hợp lệ",
     "Sản phẩm còn hàng",
     "1. Vào trang chi tiết sản phẩm\n2. Chọn size và màu\n3. Nhấn \"Mua ngay\"",
     "Sản phẩm được thêm vào giỏ và chuyển thẳng sang /thanh-toan","","","Cao","Functional"),
    (27,"TC-KH-017","Xem lịch sử đơn hàng","Khách hàng","Quản lý đơn hàng",
     "Đã đăng nhập; đã có ít nhất 1 đơn","Không có",
     "1. Đăng nhập\n2. Vào trang tài khoản → Đơn hàng của tôi",
     "Danh sách đơn hàng hiển thị đúng thứ tự và trạng thái","","","Cao","Functional"),
    (28,"TC-KH-018","Hủy đơn hàng đang chờ xử lý","Khách hàng","Quản lý đơn hàng",
     "Đã đăng nhập; có đơn trạng thái PENDING",
     "Đơn đang ở trạng thái pending",
     "1. Vào chi tiết đơn PENDING\n2. Nhấn \"Hủy đơn\"\n3. Xác nhận",
     "Đơn hàng được hủy, trạng thái chuyển thành CANCELLED","","","Cao","Functional"),
    (29,"TC-KH-019","Hủy đơn hàng đã giao thành công","Khách hàng","Quản lý đơn hàng",
     "Đã đăng nhập; có đơn trạng thái DELIVERED",
     "Đơn đang ở trạng thái delivered",
     "1. Vào chi tiết đơn DELIVERED\n2. Tìm nút hủy đơn",
     "Nút hủy đơn không hiển thị hoặc bị vô hiệu hóa","","","Cao","Functional"),
    (30,"TC-KH-020","Gửi đánh giá sản phẩm sau khi đơn đã giao","Khách hàng","Đánh giá",
     "Đã đăng nhập; đơn DELIVERED; chưa đánh giá sản phẩm này",
     "Rating: 5 sao\nNội dung: \"Sản phẩm rất đẹp, đúng mô tả\"",
     "1. Vào chi tiết đơn DELIVERED\n2. Nhập đánh giá 5 sao\n3. Nhập nội dung\n4. Nhấn \"Gửi đánh giá\"",
     "Đánh giá lưu thành công, hiển thị \"Đã gửi đánh giá thành công\"","","","Cao","Functional"),
    (31,"TC-KH-021","Gửi đánh giá khi đơn chưa giao","Khách hàng","Đánh giá",
     "Đã đăng nhập; đơn ở trạng thái PROCESSING",
     "Đơn đang xử lý",
     "1. Vào chi tiết đơn PROCESSING\n2. Kiểm tra form đánh giá",
     "Form đánh giá không hiển thị (chỉ hiện sau khi DELIVERED)","","","Cao","Functional"),
    (32,"TC-KH-022","Gửi đánh giá lần 2 cho cùng sản phẩm","Khách hàng","Đánh giá",
     "Đã đăng nhập; đã đánh giá sản phẩm trong đơn hàng này",
     "Rating: 4 sao",
     "1. Vào đơn DELIVERED đã đánh giá\n2. Kiểm tra form đánh giá",
     "Form không hiển thị hoặc hiển thị trạng thái \"Đã đánh giá\"","","","Trung bình","Functional"),
    (33,"TC-KH-023","Thêm sản phẩm vào danh sách yêu thích","Khách hàng","Wishlist",
     "Đã đăng nhập","Sản phẩm bất kỳ",
     "1. Đăng nhập\n2. Vào trang sản phẩm\n3. Nhấn icon tim / nút yêu thích",
     "Sản phẩm được thêm vào wishlist, icon đổi trạng thái","","","Trung bình","Functional"),
    (34,"TC-KH-024","Xóa sản phẩm khỏi wishlist","Khách hàng","Wishlist",
     "Đã đăng nhập; wishlist có sản phẩm","Không có",
     "1. Vào trang wishlist\n2. Nhấn xóa sản phẩm",
     "Sản phẩm bị xóa khỏi wishlist","","","Trung bình","Functional"),
    (35,"TC-KH-025","Tạo yêu cầu hỗ trợ","Khách hàng","Hỗ trợ",
     "Đã đăng nhập",
     "Nội dung: \"Tôi muốn đổi size đơn #123\"",
     "1. Đăng nhập\n2. Vào trang tài khoản\n3. Tạo ticket hỗ trợ mới\n4. Nhập nội dung\n5. Gửi",
     "Ticket được tạo, hiển thị trong danh sách yêu cầu hỗ trợ","","","Trung bình","Functional"),
    (36,"TC-KH-026","Cập nhật địa chỉ giao hàng mặc định","Khách hàng","Địa chỉ",
     "Đã đăng nhập",
     "Địa chỉ: 123 Lê Lợi, Q1, TP.HCM",
     "1. Đăng nhập\n2. Vào tài khoản → Địa chỉ\n3. Thêm địa chỉ mới\n4. Đặt làm mặc định",
     "Địa chỉ mới được lưu và đặt làm mặc định","","","Trung bình","Functional"),
    (37,"TC-KH-027","Đăng xuất khỏi hệ thống","Khách hàng","Xác thực",
     "Đã đăng nhập","Không có",
     "1. Nhấn nút Đăng xuất",
     "Phiên đăng nhập kết thúc, chuyển về /login; không thể truy cập trang yêu cầu xác thực","","","Cao","Functional"),
    (38,"TC-KH-028","Lọc sản phẩm theo danh mục Nam","Khách hàng","Tìm kiếm & lọc",
     "Chưa hoặc đã đăng nhập","Không có",
     "1. Nhấn menu Nam trên thanh điều hướng",
     "Chỉ hiển thị sản phẩm thuộc danh mục Nam","","","Trung bình","Functional"),
    (39,"TC-KH-029","Lọc sản phẩm theo khoảng giá","Khách hàng","Tìm kiếm & lọc",
     "Chưa hoặc đã đăng nhập",
     "Giá từ: 100.000đ | Giá đến: 500.000đ",
     "1. Vào /san-pham\n2. Nhập khoảng giá\n3. Áp dụng bộ lọc",
     "Chỉ hiển thị sản phẩm trong khoảng giá đã chọn","","","Trung bình","Functional"),
    (40,"TC-KH-030","Xem hồ sơ seller từ trang sản phẩm","Khách hàng","Xem seller",
     "Chưa hoặc đã đăng nhập",
     "Trang sản phẩm của 1 seller",
     "1. Vào trang chi tiết sản phẩm\n2. Nhấn tên seller",
     "Trang hồ sơ seller hiển thị thông tin cửa hàng và sản phẩm","","","Thấp","Functional"),

    # ── Người bán ───────────────────────────────────────────────────────────
    (41,"TC-NB-001","Đăng ký tài khoản Seller","Người bán","Đăng ký Seller",
     "Chưa có tài khoản seller",
     "Email: seller01@gmail.com\nTên CH: Shop Thời Trang A\nĐịa chỉ: 456 Nguyễn Huệ",
     "1. Truy cập /dang-ky-nguoi-ban\n2. Điền đầy đủ thông tin seller\n3. Nhấn Đăng ký",
     "Tài khoản seller được tạo, chuyển hướng về trang đăng nhập","","","Cao","Functional"),
    (42,"TC-NB-002","Tạo sản phẩm mới hợp lệ","Người bán","Quản lý sản phẩm",
     "Đã đăng nhập với role SELLER",
     "Tên: Áo sơ mi trắng nam\nGiá: 350.000đ\nDanh mục: Nam\nSize: M, L, XL",
     "1. Đăng nhập seller\n2. Vào Seller Center → Sản phẩm\n3. Nhấn Thêm sản phẩm\n4. Điền đầy đủ\n5. Lưu",
     "Sản phẩm được tạo thành công, hiển thị trong danh sách","","","Cao","Functional"),
    (43,"TC-NB-003","Tạo sản phẩm thiếu tên","Người bán","Quản lý sản phẩm",
     "Đã đăng nhập với role SELLER",
     "Tên: (để trống)\nGiá: 200.000đ",
     "1. Đăng nhập seller\n2. Vào Thêm sản phẩm\n3. Để trống trường tên\n4. Lưu",
     "Hệ thống hiển thị lỗi validation: tên sản phẩm là bắt buộc","","","Cao","Functional"),
    (44,"TC-NB-004","Cập nhật giá sản phẩm","Người bán","Quản lý sản phẩm",
     "Đã đăng nhập SELLER; đã có sản phẩm",
     "Giá mới: 450.000đ",
     "1. Vào danh sách sản phẩm\n2. Chọn sản phẩm cần sửa\n3. Cập nhật giá thành 450.000đ\n4. Lưu",
     "Giá sản phẩm được cập nhật và hiển thị đúng trên trang sản phẩm","","","Cao","Functional"),
    (45,"TC-NB-005","Xóa sản phẩm của mình","Người bán","Quản lý sản phẩm",
     "Đã đăng nhập SELLER; có sản phẩm",
     "Sản phẩm của chính seller",
     "1. Vào danh sách sản phẩm\n2. Chọn sản phẩm\n3. Nhấn Xóa\n4. Xác nhận",
     "Sản phẩm bị xóa, không còn hiển thị trên trang sản phẩm","","","Trung bình","Functional"),
    (46,"TC-NB-006","Thêm biến thể sản phẩm (size/màu)","Người bán","Quản lý sản phẩm",
     "Đã đăng nhập SELLER; có sản phẩm",
     "Size: XL | Màu: Đen\nGiá: 370.000đ | Tồn kho: 20",
     "1. Vào chi tiết sản phẩm\n2. Thêm biến thể mới\n3. Nhập thông tin size XL màu Đen\n4. Lưu",
     "Biến thể mới được thêm, hiển thị trên trang sản phẩm","","","Cao","Functional"),
    (47,"TC-NB-007","Tạo mã giảm giá mới","Người bán","Quản lý voucher",
     "Đã đăng nhập SELLER",
     "Mã: SELLER10\nLoại: Phần trăm\nGiá trị: 10%\nĐơn tối thiểu: 300.000đ\nHết hạn: 31/12/2025",
     "1. Vào Seller Center → Mã giảm giá\n2. Nhấn Tạo voucher\n3. Điền thông tin\n4. Lưu",
     "Voucher được tạo thành công, hiển thị trong danh sách","","","Cao","Functional"),
    (48,"TC-NB-008","Tạo mã giảm giá trùng mã đã tồn tại","Người bán","Quản lý voucher",
     "Đã đăng nhập SELLER; mã SALE10 đã tồn tại",
     "Mã: SALE10",
     "1. Vào Mã giảm giá\n2. Tạo mới với mã SALE10\n3. Lưu",
     "Hệ thống báo lỗi: \"Mã voucher 'SALE10' đã tồn tại\"","","","Cao","Functional"),
    (49,"TC-NB-009","Vô hiệu hóa mã giảm giá","Người bán","Quản lý voucher",
     "Đã đăng nhập SELLER; có voucher đang hoạt động",
     "Voucher đang active",
     "1. Vào danh sách Mã giảm giá\n2. Chọn voucher\n3. Sửa trạng thái thành Tắt\n4. Lưu",
     "Voucher bị vô hiệu hóa, khách hàng không thể áp dụng","","","Trung bình","Functional"),
    (50,"TC-NB-010","Seller không thể sửa voucher của seller khác","Người bán","Phân quyền",
     "Đã đăng nhập SELLER A; voucher của SELLER B",
     "ID voucher của seller B",
     "1. Gọi API PUT /api/seller/vouchers/{id} với id của seller khác",
     "Hệ thống trả về 403 Forbidden","","","Cao","Security"),
    (51,"TC-NB-011","Xem đơn hàng của cửa hàng","Người bán","Quản lý đơn hàng",
     "Đã đăng nhập SELLER","Không có",
     "1. Đăng nhập seller\n2. Vào Seller Center → Đơn hàng",
     "Danh sách đơn hàng của seller hiển thị đúng, có thể lọc theo trạng thái","","","Cao","Functional"),
    (52,"TC-NB-012","Cập nhật trạng thái đơn hàng","Người bán","Quản lý đơn hàng",
     "Đã đăng nhập SELLER; có đơn PENDING",
     "Trạng thái mới: CONFIRMED",
     "1. Vào chi tiết đơn PENDING\n2. Chọn trạng thái CONFIRMED\n3. Lưu",
     "Trạng thái đơn được cập nhật","","","Cao","Functional"),
    (53,"TC-NB-013","Cập nhật thông tin cửa hàng","Người bán","Hồ sơ cửa hàng",
     "Đã đăng nhập SELLER",
     "Mô tả mới: \"Chuyên áo sơ mi nam cao cấp\"",
     "1. Vào Seller Center → Tài khoản cửa hàng\n2. Cập nhật mô tả\n3. Lưu",
     "Thông tin cửa hàng được cập nhật thành công","","","Trung bình","Functional"),
    (54,"TC-NB-014","Xem báo cáo doanh thu","Người bán","Báo cáo",
     "Đã đăng nhập SELLER","Không có",
     "1. Vào Seller Center → Báo cáo",
     "Biểu đồ và số liệu doanh thu hiển thị đúng","","","Trung bình","Functional"),
    (55,"TC-NB-015","Seller không thể xóa sản phẩm của seller khác","Người bán","Phân quyền",
     "Đã đăng nhập SELLER A; sản phẩm của SELLER B",
     "ID sản phẩm của seller B",
     "1. Gọi API DELETE /api/store/products/{id} với id của seller khác",
     "Hệ thống trả về 403 Forbidden hoặc thông báo không có quyền","","","Cao","Security"),

    # ── Nhân viên kho ───────────────────────────────────────────────────────
    (56,"TC-NVK-001","Đăng nhập với tài khoản Nhân viên kho","Nhân viên kho","Đăng nhập",
     "Tài khoản WAREHOUSE đã được tạo bởi Admin",
     "Email: warehouse@mocmam.com\nMật khẩu: đúng",
     "1. Truy cập /login\n2. Nhập thông tin nhân viên kho\n3. Đăng nhập",
     "Đăng nhập thành công, chuyển đến giao diện Operations Workspace","","","Cao","Functional"),
    (57,"TC-NVK-002","Xem danh sách đơn hàng cần xử lý","Nhân viên kho","Xử lý đơn hàng",
     "Đã đăng nhập WAREHOUSE","Không có",
     "1. Đăng nhập nhân viên kho\n2. Vào Staff → Đơn hàng",
     "Danh sách đơn hàng hiển thị đầy đủ với trạng thái và thông tin xử lý","","","Cao","Functional"),
    (58,"TC-NVK-003","Cập nhật trạng thái đóng gói đơn hàng","Nhân viên kho","Xử lý đơn hàng",
     "Đã đăng nhập WAREHOUSE; đơn đang processing",
     "Đơn hàng cần đóng gói",
     "1. Vào chi tiết đơn hàng\n2. Cập nhật work state: đang đóng gói\n3. Lưu",
     "Work state đơn hàng được cập nhật thành công","","","Cao","Functional"),
    (59,"TC-NVK-004","Nhập mã vận đơn","Nhân viên kho","Xử lý đơn hàng",
     "Đã đăng nhập WAREHOUSE; đơn đã đóng gói",
     "Mã vận đơn: VN123456789",
     "1. Vào chi tiết đơn hàng\n2. Nhập mã vận đơn\n3. Lưu",
     "Mã vận đơn được lưu, đơn cập nhật trạng thái shipping","","","Cao","Functional"),
    (60,"TC-NVK-005","Xem và xử lý ticket hỗ trợ","Nhân viên kho","Hỗ trợ khách hàng",
     "Đã đăng nhập WAREHOUSE","Không có",
     "1. Vào Staff → Support Tickets\n2. Chọn ticket cần xử lý\n3. Cập nhật trạng thái",
     "Danh sách ticket hiển thị, có thể cập nhật trạng thái và trả lời","","","Trung bình","Functional"),
    (61,"TC-NVK-006","Nhân viên kho không thể truy cập Admin","Nhân viên kho","Phân quyền",
     "Đã đăng nhập WAREHOUSE","URL /admin",
     "1. Thử truy cập /admin trực tiếp",
     "Chuyển hướng về trang thích hợp hoặc báo 403 Forbidden","","","Cao","Security"),
    (62,"TC-NVK-007","Cập nhật trạng thái yêu cầu hoàn trả","Nhân viên kho","Hoàn trả",
     "Đã đăng nhập WAREHOUSE; có yêu cầu hoàn trả",
     "Trạng thái mới: Đã xử lý",
     "1. Vào danh sách hoàn trả\n2. Chọn yêu cầu\n3. Cập nhật trạng thái\n4. Lưu",
     "Trạng thái hoàn trả được cập nhật thành công","","","Trung bình","Functional"),
    (63,"TC-NVK-008","Thêm ghi chú nội bộ cho đơn hàng","Nhân viên kho","Xử lý đơn hàng",
     "Đã đăng nhập WAREHOUSE; có đơn hàng",
     "Ghi chú: \"Gói hàng dễ vỡ, cẩn thận\"",
     "1. Vào chi tiết đơn hàng\n2. Thêm ghi chú nội bộ\n3. Lưu",
     "Ghi chú được lưu thành công","","","Thấp","Functional"),

    # ── Quản trị viên ───────────────────────────────────────────────────────
    (64,"TC-QTV-001","Đăng nhập với tài khoản Admin","Quản trị viên","Đăng nhập",
     "Tài khoản ADMIN tồn tại",
     "Email: admin@mocmam.com\nMật khẩu: đúng",
     "1. Truy cập /login\n2. Nhập thông tin admin\n3. Đăng nhập",
     "Đăng nhập thành công, chuyển đến Admin Control Panel","","","Cao","Functional"),
    (65,"TC-QTV-002","Xem tổng quan hệ thống","Quản trị viên","Dashboard",
     "Đã đăng nhập ADMIN","Không có",
     "1. Đăng nhập admin\n2. Vào trang /admin",
     "Dashboard hiển thị: tổng user, đơn hàng, doanh thu, số yêu cầu pending","","","Cao","Functional"),
    (66,"TC-QTV-003","Phê duyệt yêu cầu trở thành Seller","Quản trị viên","Quản lý người dùng",
     "Đã đăng nhập ADMIN; có yêu cầu pending",
     "Yêu cầu của seller ứng viên",
     "1. Vào Admin → Phân quyền\n2. Chọn yêu cầu pending\n3. Nhấn Phê duyệt",
     "Yêu cầu được chấp thuận, role user chuyển thành SELLER","","","Cao","Functional"),
    (67,"TC-QTV-004","Từ chối yêu cầu trở thành Seller","Quản trị viên","Quản lý người dùng",
     "Đã đăng nhập ADMIN; có yêu cầu pending",
     "Yêu cầu của seller ứng viên",
     "1. Vào Admin → Phân quyền\n2. Chọn yêu cầu pending\n3. Nhấn Từ chối",
     "Yêu cầu bị từ chối, người dùng vẫn giữ role USER","","","Cao","Functional"),
    (68,"TC-QTV-005","Tạo tài khoản Nhân viên kho","Quản trị viên","Quản lý nhân viên",
     "Đã đăng nhập ADMIN",
     "Email: staff02@mocmam.com\nRole: WAREHOUSE\nHọ tên: Trần Văn B",
     "1. Vào Admin → Quản lý Staff\n2. Nhấn Tạo tài khoản\n3. Điền thông tin\n4. Lưu",
     "Tài khoản nhân viên được tạo với role WAREHOUSE","","","Cao","Functional"),
    (69,"TC-QTV-006","Khóa tài khoản người dùng vi phạm","Quản trị viên","Quản lý người dùng",
     "Đã đăng nhập ADMIN; có tài khoản cần khóa",
     "Email của user vi phạm",
     "1. Vào Admin → Quản lý người dùng\n2. Tìm user\n3. Nhấn Flag/Khóa tài khoản",
     "Tài khoản bị khóa, user không thể đăng nhập","","","Cao","Functional"),
    (70,"TC-QTV-007","Xác nhận thanh toán chuyển khoản","Quản trị viên","Quản lý đơn hàng",
     "Đã đăng nhập ADMIN; đơn chuyển khoản chưa xác nhận",
     "ID đơn hàng",
     "1. Vào Admin → Giám sát đơn hàng\n2. Tìm đơn chưa xác nhận\n3. Nhấn Xác nhận thanh toán",
     "Trạng thái thanh toán chuyển sang PAID","","","Cao","Functional"),
    (71,"TC-QTV-008","Hoàn tiền đơn hàng","Quản trị viên","Quản lý đơn hàng",
     "Đã đăng nhập ADMIN; đơn cần hoàn tiền",
     "Lý do: Sản phẩm bị lỗi",
     "1. Vào Admin → Hoàn tiền\n2. Chọn đơn hàng\n3. Nhập lý do\n4. Xác nhận",
     "Đơn hàng được cập nhật trạng thái hoàn tiền","","","Cao","Functional"),
    (72,"TC-QTV-009","Tạo đơn hàng thủ công","Quản trị viên","Tạo đơn thủ công",
     "Đã đăng nhập ADMIN",
     "Khách hàng, sản phẩm, địa chỉ đầy đủ",
     "1. Vào Admin → Tạo đơn thủ công\n2. Chọn khách hàng\n3. Thêm sản phẩm\n4. Điền địa chỉ\n5. Lưu",
     "Đơn hàng thủ công được tạo thành công","","","Trung bình","Functional"),
    (73,"TC-QTV-010","Tạo danh mục sản phẩm mới","Quản trị viên","Quản lý danh mục",
     "Đã đăng nhập ADMIN",
     "Tên: Đầm dự tiệc\nSlug: dam-du-tiec\nGiới tính: WOMEN",
     "1. Vào Admin → Danh mục & cấu hình\n2. Tạo danh mục mới\n3. Điền thông tin\n4. Lưu",
     "Danh mục được tạo và hiển thị trong menu sản phẩm","","","Trung bình","Functional"),
    (74,"TC-QTV-011","Tạo mã giảm giá toàn hệ thống","Quản trị viên","Quản lý voucher",
     "Đã đăng nhập ADMIN",
     "Mã: FLASHSALE50\nLoại: FIXED\nGiá trị: 50.000đ\nHạn: 31/01/2026",
     "1. Vào Admin → Mã giảm giá\n2. Tạo voucher mới\n3. Điền thông tin\n4. Lưu",
     "Voucher hệ thống được tạo, mọi khách hàng có thể sử dụng","","","Cao","Functional"),
    (75,"TC-QTV-012","Cập nhật cài đặt hệ thống","Quản trị viên","Cấu hình hệ thống",
     "Đã đăng nhập ADMIN",
     "orderAutoCancelHours: 24\nmaxRefundDays: 14",
     "1. Vào Admin → Danh mục & cấu hình\n2. Cập nhật thời gian tự hủy đơn = 24h\n3. Lưu",
     "Cài đặt được lưu thành công và áp dụng ngay","","","Trung bình","Functional"),
    (76,"TC-QTV-013","Xem báo cáo doanh thu theo ngày","Quản trị viên","Báo cáo",
     "Đã đăng nhập ADMIN","Không có",
     "1. Vào Admin → Báo cáo\n2. Chọn khoảng 7 ngày gần nhất",
     "Biểu đồ doanh thu theo ngày hiển thị đúng dữ liệu","","","Trung bình","Functional"),
    (77,"TC-QTV-014","Admin xem toàn bộ voucher hệ thống","Quản trị viên","Quản lý voucher",
     "Đã đăng nhập ADMIN","Không có",
     "1. Vào Admin → Mã giảm giá",
     "Danh sách tất cả voucher của admin và seller hiển thị đầy đủ","","","Trung bình","Functional"),
    (78,"TC-QTV-015","Vô hiệu hóa danh mục sản phẩm","Quản trị viên","Quản lý danh mục",
     "Đã đăng nhập ADMIN; có danh mục đang hoạt động",
     "Danh mục cần ẩn",
     "1. Vào Admin → Danh mục & cấu hình\n2. Tắt trạng thái active của danh mục",
     "Danh mục không còn hiển thị trên trang sản phẩm của khách","","","Trung bình","Functional"),
    (79,"TC-QTV-016","Xem nhật ký hệ thống","Quản trị viên","Nhật ký",
     "Đã đăng nhập ADMIN","Không có",
     "1. Vào Admin → Nhật ký hệ thống",
     "Danh sách nhật ký hiển thị các hoạt động theo thứ tự thời gian","","","Thấp","Functional"),
    (80,"TC-QTV-017","User thường không thể gọi API Admin","Quản trị viên","Phân quyền",
     "Tài khoản User thường đăng nhập",
     "Gọi API PUT /api/admin/users/{id}/role",
     "1. Đăng nhập user thường\n2. Gọi API thay đổi role",
     "Hệ thống trả về 403 Forbidden, không cho phép","","","Cao","Security"),
]

# ── sheets ───────────────────────────────────────────────────────────────────
GROUPS = [
    ("Tất cả Test Cases", list(range(len(ROWS)))),
    ("UD - Chưa đăng nhập",   [i for i,r in enumerate(ROWS) if "TC-UD" in r[1]]),
    ("KH - Khách hàng",       [i for i,r in enumerate(ROWS) if "TC-KH" in r[1]]),
    ("NB - Người bán",        [i for i,r in enumerate(ROWS) if "TC-NB" in r[1]]),
    ("NVK - Nhân viên kho",   [i for i,r in enumerate(ROWS) if "TC-NVK" in r[1]]),
    ("QTV - Quản trị viên",   [i for i,r in enumerate(ROWS) if "TC-QTV" in r[1]]),
]

COL_WIDTHS = [5, 12, 35, 22, 22, 35, 30, 50, 45, 25, 12, 14, 12]

def build_sheet(ws, indices):
    ws.freeze_panes = "A2"

    # header
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill  = PatternFill("solid", fgColor=C_HEADER)
        c.font  = Font(bold=True, color=C_HEADER_FONT, size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[1].height = 30

    actor_row_count = {}
    for data_idx in indices:
        r = ROWS[data_idx]
        actor = r[3]
        actor_row_count[actor] = actor_row_count.get(actor, 0) + 1

    actor_idx_within = {}
    for excel_row, data_idx in enumerate(indices, 2):
        r = ROWS[data_idx]
        actor = r[3]
        actor_idx_within[actor] = actor_idx_within.get(actor, 0) + 1
        colors = ACTOR_COLORS.get(actor, ("FFFFFF", "F5F5F5"))
        fill_hex = colors[0] if actor_idx_within[actor] % 2 == 1 else colors[1]

        for col, val in enumerate(r, 1):
            bold = col in (1, 2)
            align_h = "center" if col in (1, 11, 12, 13) else "left"
            font_c = "000000"
            if col == 12:  # priority
                if val == "Cao":    font_c = C_PRIORITY_H
                elif val == "Trung bình": font_c = C_PRIORITY_M
                else:               font_c = C_PRIORITY_L
            cell_style(ws, excel_row, col, val, fill_hex,
                       bold=bold, align_h=align_h, font_color=font_c)
        ws.row_dimensions[excel_row].height = 80

    # column widths
    for col, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # auto-filter
    ws.auto_filter.ref = ws.dimensions

# ── build all sheets ─────────────────────────────────────────────────────────
ws_first = True
for sheet_name, indices in GROUPS:
    ws = wb.active if ws_first else wb.create_sheet(sheet_name)
    if ws_first:
        ws.title = sheet_name
        ws_first = False
    build_sheet(ws, indices)

# ── legend sheet ─────────────────────────────────────────────────────────────
ws_leg = wb.create_sheet("Chú giải")
legend = [
    ("Actor", "Mã TC", "Màu nền", "Mô tả"),
    ("Người dùng chưa đăng nhập", "TC-UD-xxx", "Xanh dương nhạt", "User không có tài khoản"),
    ("Khách hàng",                "TC-KH-xxx", "Xanh lá nhạt",   "User đã đăng ký, role USER"),
    ("Người bán",                 "TC-NB-xxx", "Vàng nhạt",      "User role SELLER"),
    ("Nhân viên kho",             "TC-NVK-xxx","Cam nhạt",       "User role WAREHOUSE/STYLES"),
    ("Quản trị viên",             "TC-QTV-xxx","Tím nhạt",       "User role ADMIN"),
]
actor_fills = {
    "Người dùng chưa đăng nhập": C_UD,
    "Khách hàng": C_KH,
    "Người bán": C_NB,
    "Nhân viên kho": C_NVK,
    "Quản trị viên": C_QTV,
}
for ri, row in enumerate(legend, 1):
    for ci, val in enumerate(row, 1):
        c = ws_leg.cell(row=ri, column=ci, value=val)
        if ri == 1:
            c.fill = PatternFill("solid", fgColor=C_HEADER)
            c.font = Font(bold=True, color=C_HEADER_FONT, size=10, name="Calibri")
        elif ci == 1 and val in actor_fills:
            c.fill = PatternFill("solid", fgColor=actor_fills[val])
            c.font = Font(size=9, name="Calibri")
        else:
            c.font = Font(size=9, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws_leg.row_dimensions[ri].height = 20
for ci, w in enumerate([30, 15, 20, 40], 1):
    ws_leg.column_dimensions[get_column_letter(ci)].width = w

out = r"d:\Test - Copy (1)\Nhom_6-TestCases-MocMam_ChinhThuc.xlsx"
wb.save(out)
print("OK:", out)
